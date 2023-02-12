# 目标：
#     接收两个整数和一个文件名字符串做为命令行参数
#     第一个整数N，程序从第N行开始插入
#     第二个整数M，程序插入N个空行
# 操作：
#     1.接收整数
#     2.接收文件名
#     3.打开电子表格
#     4.1 利用for循环从第'N'行起，依次将第'N'行的内容赋给第'N+M'行，直至将第'row_max'行赋值给第'row_max+M'行
#         --sheet.cell(row = rowNum + conti_m, column= j).value = sheet.cell(row = rowNum, column= j).value
#     4.2 从第'N'行起,直到第'N+M'行放空值 
#         --sheet.cell(row = rowNum, column = j).value = None
#     5.生成新表



import openpyxl
from pathlib import Path

print('Start...')
print('Please input the start line to create empty: ')
line_n = int(input()) # M

print('Input how many lines you want to create empty: ')
conti_m = int(input()) # N 
print('Better to privide the absolute address')
# open_file = input()
open_file = 'produceSales.xlsx'
# open_file = Path.cwd() / open_file # 文件夹的绝对路径要心知肚明
# print(open_file)
wb = openpyxl.load_workbook(open_file)
sheet = wb['Sheet1']

print('Running...')
# 从第'N'行起，遍历M行
for rowNum in range(line_n, line_n + conti_m):
    for j in range(1, sheet.max_column +1):
        # 将'N'行的内容依次赋给'N+M'行 
        sheet.cell(row = rowNum + conti_m, column= j).value = sheet.cell(row = rowNum, column= j).value
        # 将'N'行起的内容赋值None
        sheet.cell(row = rowNum, column = j).value = None


print('Closing...')
wb.save('addEmptyLine_'+open_file)
print('Done.')