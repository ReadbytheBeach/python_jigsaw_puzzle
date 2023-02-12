from openpyxl.styles import Font
import openpyxl



wb = openpyxl.Workbook()
sheet = wb['Sheet']
'''
创建sheet2
'''
sheet2=wb.create_sheet(index=1, title='Sheet2')

'''
创建字体格式
'''
italic24Font = Font(size= 24, italic=True)
sheet['A1'].font = italic24Font
sheet['A1'] = 'Hello World!'

# 创建一个Font的对象
fontobj = Font(name='Times New Roman', bold =True)
sheet['B1'].font = fontobj
sheet['B1'] = 'New Style'

'''
excel使用公式
'''
sheet['C9'] = '=sum(C1:C8)'


'''
修改行高列宽
'''
# 行高中的行使用数字，列宽中的列使用英文字母
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width =50
sheet.column_dimensions['D'].width = 0  # 设置为0，没什么效果


'''
合并或拆分单元格
'''
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Merge the unit'
sheet['A1'].font = italic24Font

sheet.merge_cells('D4:F4')
sheet['D4'] = 'try merging'
sheet.unmerge_cells('D4:F4')

'''
冻结窗口
'''
sheet2=wb.active
sheet2.freeze_panes = 'C2'


wb.save('styles.xlsx')


