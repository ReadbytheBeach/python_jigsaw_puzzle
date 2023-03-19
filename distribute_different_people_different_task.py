# 目标： 
# 接收一个电子邮件的列表，
# 随机将工作发给不同的组员（每个工作仅发放一次，不重复）
# 用电子邮件通知每个小组成员分配给他们的工作， 
# 安排程序每周自动运行一次  


'''
you need create an excel first
you need close the company VPN

# 接收一个电子邮件的列表，
    # 获取人名和邮件名
        wb = openpyxl.load_workbook('distribute_different people different task.xlsx')
        sheet = wb.get_sheet_by_name('Sheet1')
        last_colmn = sheet.max_column
        # 名字 + 邮件是一个二维结构，用键-值对的字典是一个好的解决方案
        members_mail = {}
        for r in range(2, sheet.max_row+1):
            name = sheet.cell(row=r, column=1).value
            mail = sheet.cell(row=r, column=last_colmn).value
            members_mail[name] = mail
# 随机将工作发给不同的组员（每个工作仅发放一次，不重复）
    # 随机选择任务并分发， 不重复   
        random_task = random.choice(tasks)
        tasks.remove(random_task)
# 用电子邮件通知每个小组成员分配给他们的工作， 
    # 创建SMTP链接
        email_from = 'Heinekenblue@163.com'
        print('input your email password here: ')
        pswd = input()
        # 使用SMTP协议链接到网络
        smtpObj = smtplib.SMTP('smtp.163.com', 25)
        # 调用ehlo()方法向SMTP电子邮件服务打招呼
        smtpObj.ehlo()
        # 使用TSL加密模式必须啊告知，如果使用SSL模式则跳过以下语句
        smtpObj.starttls()
        # 登入发件人邮箱
        smtpObj.login(email_from, pswd)
    # 写邮件主题，正文
        body 写邮件的技巧
    # 依次分发
        sendmailStatus = smtpObj.sendmail(email_from, email_to, body)
# 安排程序每周自动运行一次  
    # time_left = 60*24*7
    # while time_left > 0:
    #     time_left = time_left -1
'''

import openpyxl, smtplib, random, time, os

# 改变当前目录
os.chdir(r'D:\03_program\python\python_jigsaw_puzzle')
print(os.getcwd())


# Tasks should be equal or more than member
# Tasks 也可以以表格方式导入， 有邮件列表相同
tasks = ['data collection', 'data cleaning', 'data mining', 'data upload','data management', 'data search']
# 每周的时间 = 60 * 24 * 7 
time_left = 10 # 60*24*7
while time_left > 0:
    # 查看剩余时间
    # print('time_left = ', time_left)
    # time.sleep(1)
    time_left = time_left -1

wb = openpyxl.load_workbook('distribute_different people different task.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
last_colmn = sheet.max_column

# 名字 + 邮件是一个二维结构，用键-值对的字典是一个好的解决方案
members_mail = {}
for r in range(2, sheet.max_row+1):
    name = sheet.cell(row=r, column=1).value
    mail = sheet.cell(row=r, column=last_colmn).value
    members_mail[name] = mail


# log into email account
email_from = 'Heinekenblue@163.com'
print('input your email password here: ')
pswd = input()
# 使用SMTP协议链接到网络
smtpObj = smtplib.SMTP('smtp.163.com', 25)
# 调用ehlo()方法向SMTP电子邮件服务打招呼
smtpObj.ehlo()
# 使用TSL加密模式必须啊告知，如果使用SSL模式则跳过以下语句
smtpObj.starttls()
# 登入发件人邮箱
smtpObj.login(email_from, pswd)

# 开始发放任务
for name, email_to in members_mail.items():
    # 随机抽取任务
    random_task = random.choice(tasks)
    tasks.remove(random_task)

    # 主题和正文之间需要用空行分隔，这是SMTP的协议要求，即使用“Subject: XXX. \n\nDear, XXX”
    body = "Subject: %s Task is coming... \n\nDear %s, \n\n    Now you be assigned the task: %s.\n\nBR,\nxj"%(random_task, name, random_task)
    print('sending email to %s...'%(name))

    try:
        # 发送邮件
        sendmailStatus = smtpObj.sendmail(email_from, email_to, body)
    except:
        # 邮箱错误，即sendmailStatus != {}:
        print('There was a problem sending emial to %s: %s' %(email_to, sendmailStatus))  # 说明这个客户的邮箱已经不存在了, 你懂的

smtpObj.quit()
print('mail has been sented to all members.')
print()
print('Remind: %s tasks still has not been asigned'%(tasks))
     



