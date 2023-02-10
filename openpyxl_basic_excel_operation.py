import openpyxl,os
import logging

logging.basicConfig(level=logging.DEBUG ,format='%(asctime)s - %(levelname)s - %(message)s')
# logging.disable(logging.DEBUG)
logging.info('Start the program.')


print(os.getcwd())  # 看一下当前目录
os.chdir(r'D:\03_program\python\automate_the_boring_stuff_with_python') # 变更当前目录
logging.debug('current director: '+ str(os.getcwd()))

'''
取得工作表
'''
wb = openpyxl.load_workbook('example.xlsx')  # wb: workbook
logging.debug('workbook type: ' + str(type(wb)))
logging.debug('wb sheetname: '+ str(wb.sheetnames))

sheet = wb['Sheet3']  # 从excel中取得一个sheet, 注意python区分大小写
logging.debug('sheet type: ' + str(type(sheet)))
logging.debug('sheet.title: ' + str(sheet.title))

anotherSheet = wb.active  # 激活worksheet
logging.debug('anothersheet: ' + str(anotherSheet))

'''
取得表中的单元格
'''
sheet = wb['Sheet1']
logging.debug("Unit-B3 in Sheet1: " + str(sheet['B3']))  # 取Sheet1，的'B3'单元格
logging.debug('the value of Unit-B3 in Sheet1: '+str(sheet['B3'].value))
c = sheet['C3']  # 另一种方法取得单元格, 貌似可以不区分单元格坐标系名的大小写
logging.debug('the value of Unit-C3 in Sheet1: ' +str(c.value))
logging.debug('Row %s, Column %s, Value %s' %(c.row, c.column, c.value)) # 取单元格的行号，列号，值
logging.debug('Cell %s is %s'%(c.coordinate,c.value))  # 取单元格的坐标系名--C3

print(sheet.cell(row=3, column=2))
print(sheet.cell(row=3,column=2).value)
for i in range(1,8,2): # 每隔一行输出 列=3的单元格值
    print(i, sheet.cell(row=i, column =3).value)

'''
确定表的大小
'''
print('max row: ',sheet.max_row) # max_row, max_column都是sheet的属性
print('max column: ', sheet.max_column)

'''
切片方法，取行、列、或单元格
'''
# print(tuple(sheet['A1':'G3']))
print(sheet['A1':'G3'])  # 不用tuple()也能正确显示
for rowOfCellObjects in sheet['A1':'C3']:  # 针对切片中的每一行
    for cellObj in rowOfCellObjects:    # 针对行中的单元格
        print(cellObj.coordinate, cellObj.value)
    print('-----End Of Row------')

'''
用list方法，取行、列、单元格
'''
sheet = wb.active
# print(list(sheet.columns)[1])  # 取得第二列的所有单元
print(list(sheet.columns)[1])
print(list(sheet.rows)[1])
for cellObj in list(sheet.columns)[2]:
    print(cellObj.value)



logging.info('End the program.')