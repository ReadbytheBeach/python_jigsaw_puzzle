# 目的：
    # 循环遍历所有行
    # 如果该行是Garlic, Celery或Lemon, 就更新价格
# 操作
    # 打开电子表格
    # 逐行检查是否是Garlic, Celery or Lemon
    # 如果是，就更新相应的单价
    # 保存成一个新版的电子表格

'''
优雅的更新数据 --- 使用字典的方式


    老的方法：
        if produceName == 'Celery':
            cellObj = 1.19
        if produceName == 'Garlic':
            cellObj = 3.07
        if produceName == 'Lemon'
            cellObj = 1.27

    新方法
    （也便于以后更改）：
        PRICE_UPDATES = {'Garlic': 3.07,
                        'Celery': 1.19,
                        'Lemon': 1.27
                        }
'''

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet1']

# The produce types and they new price
PRICE_UPDATES = {'Garlic': 3.07,
                'Celery': 1.19,
                'Lemon': 1.27
                }

# Loop through the rows and update the prices
'''
逐行比较
找到第一例的单元格
比较单元格的值（菜名） ?=  更新字典的键（需更新的菜名）
如果是，单元格的值 = 更新为字典[键] 所对应的值
'''
for rowNum in range(2, sheet.max_row):  # skip the first row
    produceName = sheet.cell(row = rowNum, column = 1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row = rowNum, column = 2).value = PRICE_UPDATES[produceName]

wb.save('new_produceSales.xlsx')
