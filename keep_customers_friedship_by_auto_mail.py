# 目标： 向会员发送新一期的内容
# 从Excel电子表格中读取数据
# 找出上个月已缴费的所有会员
# 找到他们的电子邮件地址，向他们发送针对个人的新的一期宣传材料
'''
策略：
    用openpyxl模块打开读取Excel文档的单元格(见之前的程序)
    创建一个字典，包含会员已缴费的人员名单
    调用smtplib.SMTP()、ehlo().starttls()和login() 登录服务器
    针对已缴费的所有会员， 调用sendmail()方法， 发送针对个人的电子邮件提醒
'''

import openpyxl, smtplib, sys

# open the spreadsheet and get the latest dues status
wb = openpyxl.load_workbook('customer_list.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
lastCol = sheet.max_column
# print(lastCol)
latestMonth = sheet.cell(row = 1, column= lastCol).value
# print(latestMonth)

# check each member's sent mail status
unsentCustomers = {}
for r in range(2, sheet.max_row + 1):
    sent_status = sheet.cell(row= r, column=lastCol).value
    if sent_status != 'sent':
        name = sheet.cell(row= r, column= 1).value
        email = sheet.cell(row= r, column= 2).value 
        unsentCustomers[name] = email

print('latest Festiva: \"%s\" unsent mail customer name: %s'%(latestMonth, unsentCustomers.keys()))

# log into email account
email_from = 'Heinekenblue@163.com'
print('input your email password here: ')
pswd = input()
smtpObj = smtplib.SMTP('smtp.163.com', 25)
# 调用ehlo()方法向SMTP电子邮件服务打招呼
smtpObj.ehlo()
# 使用TSL加密模式必须啊告知，如果使用SSL模式则跳过以下语句
smtpObj.starttls()
# 登入发件人邮箱
smtpObj.login(email_from, pswd)

# send out friendship emails
for name, email_to in unsentCustomers.items():
    # 主题和正文之间需要用空行分隔，这是SMTP的协议要求，即使用“Subject: XXX. \n\nDear, XXX”
    body = "Subject: %s is coming, long time not see. \n\nDear %s, \n\n    Seems we have sometime not seen each other. How are you these days? \n\n    Hope to see you in the coming days. \n    Let's talk about your harvest and perceptions, and see what vision for future. \n\nBR,\nxj"%(latestMonth,name)
    print('sending email to %s...'%email_to)

    try:
        sendmailStatus = smtpObj.sendmail(email_from, email_to, body)
    except:
        # 邮箱错误，即sendmailStatus != {}:
        print('There was a problem sending emial to %s: %s' %(email_to, sendmailStatus))  # 说明这个客户的邮箱已经不存在了, 你懂的

smtpObj.quit()
print('mail has been sented to all customers.')






